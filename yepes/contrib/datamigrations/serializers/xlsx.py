# -*- coding:utf-8 -*-

from __future__ import absolute_import, unicode_literals

from openpyxl import load_workbook, Workbook

from yepes.contrib.datamigrations.serializers import Serializer
from yepes.types import Undefined


class XlsxSerializer(Serializer):

    is_binary = True

    def __init__(self, **serializer_parameters):
        serializer_parameters.setdefault('nonereplacement', '\\N')
        super(XlsxSerializer, self).__init__(**serializer_parameters)

    def dump(self, headers, data, file):
        workbook = Workbook(write_only=True)
        worksheet = workbook.create_sheet()
        worksheet.append(headers)
        for row in data:
            worksheet.append(row)

        workbook.save(file)

    def load(self, headers, file):
        workbook = load_workbook(file)
        worksheet = workbook.active
        data = worksheet.iter_rows()

        first_line = [
            cell.value
            for cell
            in next(data)
        ]
        if first_line == headers:
            return ([cell.value for cell in row] for row in data)

        positions = []
        for header in headers:
            try:
                pos = first_line.index(header)
            except ValueError:
                pos = None

            positions.append(pos)

        return (
            [
                Undefined if pos is None else row[pos].value
                for pos
                in positions
            ]
            for row
            in data
        )

